"""Phase 5.3/5.4: PDF transcript + Excel export service."""
import io
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.class_ import Class
from app.models.course import Course
from app.models.grade import Grade
from app.models.score import Score
from app.models.semester import Semester
from app.models.student import Student


# ── PDF helpers ──────────────────────────────────────────────────────────────

_TRANSCRIPT_HTML = """\
<html>
<head><meta charset="utf-8">
<style>
body {{ font-family: SimSun, "Noto Sans CJK SC", sans-serif; font-size: 12px; }}
h1 {{ text-align: center; font-size: 18px; }}
h2 {{ text-align: center; font-size: 14px; color: #555; }}
table {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
th, td {{ border: 1px solid #333; padding: 6px 8px; text-align: center; }}
th {{ background: #e0e0e0; }}
.info {{ margin: 8px 0; }}
.footer {{ margin-top: 20px; text-align: right; font-size: 10px; color: #999; }}
.fail {{ color: red; }}
</style></head>
<body>
<h1>学生成绩单</h1>
<h2>{semester_name}</h2>
<div class="info"><b>姓名：</b>{student_name} &nbsp;&nbsp; <b>学号：</b>{student_no} &nbsp;&nbsp; <b>班级：</b>{class_name}</div>
<table>
<tr><th>课程</th><th>分数</th><th>班级排名</th><th>班级均分</th></tr>
{rows}
<tr><th>总分</th><td><b>{total_score}</b></td><td colspan="2">班级排名: {class_rank} &nbsp; 年级排名: {grade_rank}</td></tr>
</table>
<div class="footer">打印日期: {print_date}</div>
</body></html>
"""


def _build_transcript_html(db: Session, student: Student, semester: Semester) -> str | None:
    """Build HTML string for a single student transcript."""
    from app.services.statistics_service import student_statistics
    from datetime import date

    stats = student_statistics(db, student.id, semester.id)
    if stats is None:
        return None

    cls = db.get(Class, student.class_id)

    rows = ""
    for subj in stats["subjects"]:
        fail_cls = ' class="fail"' if subj["score"] < 60 else ""
        rows += f'<tr><td>{subj["course_name"]}</td><td{fail_cls}>{subj["score"]}</td>'
        rows += f'<td>{subj["class_rank"]}/{subj["class_total"]}</td><td>{subj["class_avg"]}</td></tr>\n'

    return _TRANSCRIPT_HTML.format(
        semester_name=semester.name,
        student_name=student.name,
        student_no=student.student_no,
        class_name=cls.name if cls else "",
        rows=rows,
        total_score=stats["total_score"],
        class_rank=stats["total_class_rank"],
        grade_rank=stats["total_grade_rank"],
        print_date=date.today().isoformat(),
    )


def _html_to_pdf(html: str) -> bytes:
    """Convert HTML string to PDF bytes using xhtml2pdf (Windows-safe)."""
    from xhtml2pdf import pisa

    buf = io.BytesIO()
    pisa.CreatePDF(io.StringIO(html), dest=buf)
    return buf.getvalue()


def generate_transcript_pdf(db: Session, student_id: int, semester_id: int) -> bytes | None:
    student = db.get(Student, student_id)
    if not student:
        return None
    semester = db.get(Semester, semester_id)
    if not semester:
        return None

    html = _build_transcript_html(db, student, semester)
    if html is None:
        return None

    return _html_to_pdf(html)


def generate_batch_transcript_pdf(db: Session, class_id: int, semester_id: int) -> bytes | None:
    """Generate a combined PDF for all students in a class."""
    semester = db.get(Semester, semester_id)
    if not semester:
        return None

    students = db.query(Student).filter(Student.class_id == class_id).order_by(Student.student_no).all()
    if not students:
        return None

    combined_html = ""
    for stu in students:
        html = _build_transcript_html(db, stu, semester)
        if html:
            combined_html += html + '<div style="page-break-after: always;"></div>'

    if not combined_html:
        return None

    return _html_to_pdf(combined_html)


# ── Excel helpers ────────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
_STAT_FONT = Font(bold=True)
_FAIL_FONT = Font(color="FF0000")
_CENTER = Alignment(horizontal="center", vertical="center")


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val.encode("gbk", errors="replace")))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _rate(count: int, total: int) -> str:
    if total == 0:
        return "0%"
    return f"{round(count / total * 100, 1)}%"


def generate_class_summary_excel(db: Session, class_id: int, semester_id: int) -> bytes:
    """班级汇总表: 学号 + 姓名 + 各科 + 总分 + 排名, 底部统计行."""
    cls = db.get(Class, class_id)
    students = db.query(Student).filter(Student.class_id == class_id).all()
    student_map = {s.id: s for s in students}

    # Get all scores for this class/semester
    scores = (
        db.query(Score)
        .filter(Score.class_id == class_id, Score.semester_id == semester_id)
        .all()
    )

    # Determine courses
    course_ids = sorted({s.course_id for s in scores})
    courses = {c.id: c.name for c in db.query(Course).filter(Course.id.in_(course_ids)).all()}
    course_order = [cid for cid in course_ids]

    # Build per-student data
    data: dict[int, dict] = {}
    for s in scores:
        data.setdefault(s.student_id, {})[s.course_id] = float(s.score)

    # Sort by total descending
    entries = []
    for sid, score_map in data.items():
        stu = student_map.get(sid)
        if not stu:
            continue
        total = sum(score_map.values())
        entries.append({"student": stu, "scores": score_map, "total": total})
    entries.sort(key=lambda e: e["total"], reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{cls.name if cls else '班级'}成绩汇总"

    # Header
    headers = ["排名", "学号", "姓名"] + [courses.get(cid, "") for cid in course_order] + ["总分"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER

    # Data rows
    for rank, entry in enumerate(entries, 1):
        stu = entry["student"]
        row = rank + 1
        ws.cell(row=row, column=1, value=rank).alignment = _CENTER
        ws.cell(row=row, column=2, value=stu.student_no)
        ws.cell(row=row, column=3, value=stu.name)
        for ci, cid in enumerate(course_order, 4):
            val = entry["scores"].get(cid)
            cell = ws.cell(row=row, column=ci, value=val)
            if val is not None and val < 60:
                cell.font = _FAIL_FONT
        ws.cell(row=row, column=len(headers), value=entry["total"])

    # Stats rows
    stat_row = len(entries) + 2
    all_scores_by_course: dict[int, list[float]] = {}
    for entry in entries:
        for cid, val in entry["scores"].items():
            all_scores_by_course.setdefault(cid, []).append(val)

    stats_labels = ["平均分", "最高分", "最低分", "及格率", "良好率", "优秀率", "低分率", "超低分率"]
    for si, label in enumerate(stats_labels):
        r = stat_row + si
        ws.cell(row=r, column=1, value="").alignment = _CENTER
        ws.cell(row=r, column=2, value="")
        cell = ws.cell(row=r, column=3, value=label)
        cell.font = _STAT_FONT
        for ci, cid in enumerate(course_order, 4):
            vals = all_scores_by_course.get(cid, [])
            if not vals:
                continue
            n = len(vals)
            if label == "平均分":
                ws.cell(row=r, column=ci, value=round(sum(vals) / n, 1))
            elif label == "最高分":
                ws.cell(row=r, column=ci, value=max(vals))
            elif label == "最低分":
                ws.cell(row=r, column=ci, value=min(vals))
            elif label == "及格率":
                ws.cell(row=r, column=ci, value=_rate(sum(1 for v in vals if v >= 60), n))
            elif label == "良好率":
                ws.cell(row=r, column=ci, value=_rate(sum(1 for v in vals if v >= 70), n))
            elif label == "优秀率":
                ws.cell(row=r, column=ci, value=_rate(sum(1 for v in vals if v >= 85), n))
            elif label == "低分率":
                ws.cell(row=r, column=ci, value=_rate(sum(1 for v in vals if v < 40), n))
            elif label == "超低分率":
                ws.cell(row=r, column=ci, value=_rate(sum(1 for v in vals if v < 30), n))

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_grade_ranking_excel(db: Session, grade_id: int, semester_id: int) -> bytes:
    """年级排名表: 排名 + 学号 + 姓名 + 班级 + 各科 + 总分."""
    classes = db.query(Class).filter(Class.grade_id == grade_id).all()
    class_ids = [c.id for c in classes]
    class_map = {c.id: c.name for c in classes}

    students = db.query(Student).filter(Student.class_id.in_(class_ids)).all()
    student_map = {s.id: s for s in students}

    scores = (
        db.query(Score)
        .filter(Score.class_id.in_(class_ids), Score.semester_id == semester_id)
        .all()
    )

    course_ids = sorted({s.course_id for s in scores})
    courses = {c.id: c.name for c in db.query(Course).filter(Course.id.in_(course_ids)).all()}

    data: dict[int, dict] = {}
    for s in scores:
        data.setdefault(s.student_id, {})[s.course_id] = float(s.score)

    entries = []
    for sid, score_map in data.items():
        stu = student_map.get(sid)
        if not stu:
            continue
        entries.append({"student": stu, "scores": score_map, "total": sum(score_map.values())})
    entries.sort(key=lambda e: e["total"], reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "年级排名"

    headers = ["排名", "学号", "姓名", "班级"] + [courses.get(cid, "") for cid in course_ids] + ["总分"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER

    for rank, entry in enumerate(entries, 1):
        stu = entry["student"]
        row = rank + 1
        ws.cell(row=row, column=1, value=rank).alignment = _CENTER
        ws.cell(row=row, column=2, value=stu.student_no)
        ws.cell(row=row, column=3, value=stu.name)
        ws.cell(row=row, column=4, value=class_map.get(stu.class_id, ""))
        for ci, cid in enumerate(course_ids, 5):
            val = entry["scores"].get(cid)
            cell = ws.cell(row=row, column=ci, value=val)
            if val is not None and val < 60:
                cell.font = _FAIL_FONT
        ws.cell(row=row, column=len(headers), value=entry["total"])

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_student_scores_excel(db: Session, student_id: int) -> bytes:
    """学生多学期成绩: 每学期一个 Sheet."""
    student = db.get(Student, student_id)

    # Find all semesters with scores
    semesters = (
        db.query(Semester)
        .join(Score, Score.semester_id == Semester.id)
        .filter(Score.student_id == student_id)
        .distinct()
        .order_by(Semester.start_date)
        .all()
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for sem in semesters:
        ws = wb.create_sheet(title=sem.name[:31])  # Excel sheet name max 31 chars

        scores = (
            db.query(Score)
            .filter(Score.student_id == student_id, Score.semester_id == sem.id)
            .all()
        )
        course_ids = [s.course_id for s in scores]
        courses = {c.id: c.name for c in db.query(Course).filter(Course.id.in_(course_ids)).all()}

        headers = ["课程", "分数"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER

        for ri, sc in enumerate(scores, 2):
            ws.cell(row=ri, column=1, value=courses.get(sc.course_id, ""))
            cell = ws.cell(row=ri, column=2, value=float(sc.score))
            if float(sc.score) < 60:
                cell.font = _FAIL_FONT

        total = sum(float(s.score) for s in scores)
        r = len(scores) + 2
        ws.cell(row=r, column=1, value="总分").font = _STAT_FONT
        ws.cell(row=r, column=2, value=total).font = _STAT_FONT

        _auto_width(ws)

    if not wb.sheetnames:
        wb.create_sheet("空")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
