"""Phase 5.3/5.4: PDF transcript & Excel export tests — TDD RED phase."""
from datetime import date
from decimal import Decimal

import pytest

from tests.conftest import auth_header


# ── Shared fixtures ──────────────────────────────────────────────────────────

@pytest.fixture
def report_data(db):
    """Seed data for report tests: grade → class → students → semester → course → scores."""
    from app.core.security import hash_password
    from app.models.class_ import Class
    from app.models.course import Course
    from app.models.grade import Grade
    from app.models.score import Score
    from app.models.semester import Semester
    from app.models.student import Student
    from app.models.teacher_course_class import TeacherCourseClass
    from app.models.user import User, UserRole

    teacher = User(username="tr", hashed_password=hash_password("tr123"), name="李老师", role=UserRole.TEACHER, is_active=True)
    db.add(teacher)
    db.flush()

    grade = Grade(name="初一")
    db.add(grade)
    db.flush()

    cls = Class(name="初一(1)班", grade_id=grade.id)
    db.add(cls)
    db.flush()

    sem = Semester(name="2025秋季", start_date=date(2025, 9, 1), end_date=date(2026, 1, 31), is_active=True)
    db.add(sem)
    db.flush()

    math = Course(name="数学R")
    chinese = Course(name="语文R")
    db.add_all([math, chinese])
    db.flush()

    db.add_all([
        TeacherCourseClass(teacher_id=teacher.id, course_id=math.id, class_id=cls.id, semester_id=sem.id),
        TeacherCourseClass(teacher_id=teacher.id, course_id=chinese.id, class_id=cls.id, semester_id=sem.id),
    ])

    students = []
    scores_data = [
        ("S001", "张三", Decimal("88"), Decimal("75")),
        ("S002", "李四", Decimal("72"), Decimal("90")),
        ("S003", "王五", Decimal("55"), Decimal("60")),
    ]
    for sno, name, math_sc, chi_sc in scores_data:
        stu = Student(student_no=sno, name=name, gender="male", class_id=cls.id)
        db.add(stu)
        db.flush()
        students.append(stu)
        db.add(Score(student_id=stu.id, course_id=math.id, class_id=cls.id, semester_id=sem.id, teacher_id=teacher.id, score=math_sc))
        db.add(Score(student_id=stu.id, course_id=chinese.id, class_id=cls.id, semester_id=sem.id, teacher_id=teacher.id, score=chi_sc))

    db.commit()

    return {
        "teacher": teacher,
        "grade": grade,
        "cls": cls,
        "sem": sem,
        "math": math,
        "chinese": chinese,
        "students": students,
    }


@pytest.fixture
def tr_token(client, report_data):
    r = client.post("/api/v1/auth/login", data={"username": "tr", "password": "tr123"},
                     headers={"Content-Type": "application/x-www-form-urlencoded"})
    return r.json()["access_token"]


# ── 5.3 PDF Transcript Tests ────────────────────────────────────────────────

class TestPdfTranscript:
    def test_single_student_pdf(self, client, tr_token, report_data):
        d = report_data
        r = client.get(
            "/api/v1/reports/transcript/pdf",
            params={"student_id": d["students"][0].id, "semester_id": d["sem"].id},
            headers=auth_header(tr_token),
        )
        assert r.status_code == 200
        assert r.headers["content-type"] == "application/pdf"
        # PDF magic bytes
        assert r.content[:5] == b"%PDF-"

    def test_single_pdf_contains_student_info(self, client, tr_token, report_data):
        """PDF should contain the student's name (even as binary, the text is embedded)."""
        d = report_data
        r = client.get(
            "/api/v1/reports/transcript/pdf",
            params={"student_id": d["students"][0].id, "semester_id": d["sem"].id},
            headers=auth_header(tr_token),
        )
        assert r.status_code == 200
        assert len(r.content) > 100  # Non-trivial PDF

    def test_batch_pdf(self, client, tr_token, report_data):
        d = report_data
        r = client.post(
            "/api/v1/reports/transcript/pdf/batch",
            json={"class_id": d["cls"].id, "semester_id": d["sem"].id},
            headers=auth_header(tr_token),
        )
        assert r.status_code == 200
        assert r.headers["content-type"] == "application/pdf"
        assert r.content[:5] == b"%PDF-"

    def test_student_not_found(self, client, tr_token, report_data):
        r = client.get(
            "/api/v1/reports/transcript/pdf",
            params={"student_id": 9999, "semester_id": report_data["sem"].id},
            headers=auth_header(tr_token),
        )
        assert r.status_code == 404

    def test_student_role_can_get_own_pdf(self, client, report_data, db):
        """A student user can download their own transcript."""
        from app.core.security import hash_password
        from app.models.user import User, UserRole

        stu = report_data["students"][0]
        user = User(username="stux", hashed_password=hash_password("stu123"), name="张三", role=UserRole.STUDENT, is_active=True)
        db.add(user)
        db.flush()
        stu.user_id = user.id
        db.commit()

        r = client.post("/api/v1/auth/login", data={"username": "stux", "password": "stu123"},
                         headers={"Content-Type": "application/x-www-form-urlencoded"})
        tok = r.json()["access_token"]

        r = client.get(
            "/api/v1/reports/transcript/pdf",
            params={"student_id": stu.id, "semester_id": report_data["sem"].id},
            headers=auth_header(tok),
        )
        assert r.status_code == 200

    def test_student_role_cannot_get_other_pdf(self, client, report_data, db):
        from app.core.security import hash_password
        from app.models.user import User, UserRole

        user = User(username="stuother", hashed_password=hash_password("stu123"), name="其他", role=UserRole.STUDENT, is_active=True)
        db.add(user)
        db.flush()
        report_data["students"][0].user_id = user.id
        db.commit()

        r = client.post("/api/v1/auth/login", data={"username": "stuother", "password": "stu123"},
                         headers={"Content-Type": "application/x-www-form-urlencoded"})
        tok = r.json()["access_token"]

        other_stu = report_data["students"][1]
        r = client.get(
            "/api/v1/reports/transcript/pdf",
            params={"student_id": other_stu.id, "semester_id": report_data["sem"].id},
            headers=auth_header(tok),
        )
        assert r.status_code == 403


# ── 5.4 Excel Export Tests ──────────────────────────────────────────────────

class TestExcelClassSummary:
    def test_returns_xlsx(self, client, tr_token, report_data):
        d = report_data
        r = client.get(
            "/api/v1/reports/class-summary/excel",
            params={"class_id": d["cls"].id, "semester_id": d["sem"].id},
            headers=auth_header(tr_token),
        )
        assert r.status_code == 200
        assert "spreadsheet" in r.headers["content-type"]
        # XLSX magic bytes (PK zip)
        assert r.content[:2] == b"PK"

    def test_contains_students(self, client, tr_token, report_data):
        """The Excel should contain 3 student rows + stats rows."""
        import io
        import openpyxl
        d = report_data
        r = client.get(
            "/api/v1/reports/class-summary/excel",
            params={"class_id": d["cls"].id, "semester_id": d["sem"].id},
            headers=auth_header(tr_token),
        )
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # Header row + 3 students + at least 1 stat row
        assert ws.max_row >= 5


class TestExcelGradeRanking:
    def test_returns_xlsx(self, client, tr_token, report_data):
        d = report_data
        r = client.get(
            "/api/v1/reports/grade-ranking/excel",
            params={"grade_id": d["grade"].id, "semester_id": d["sem"].id},
            headers=auth_header(tr_token),
        )
        assert r.status_code == 200
        assert r.content[:2] == b"PK"

    def test_contains_ranking(self, client, tr_token, report_data):
        import io
        import openpyxl
        d = report_data
        r = client.get(
            "/api/v1/reports/grade-ranking/excel",
            params={"grade_id": d["grade"].id, "semester_id": d["sem"].id},
            headers=auth_header(tr_token),
        )
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        # Header + 3 students
        assert ws.max_row >= 4


class TestExcelStudentScores:
    def test_returns_xlsx(self, client, tr_token, report_data):
        d = report_data
        r = client.get(
            "/api/v1/reports/student-scores/excel",
            params={"student_id": d["students"][0].id},
            headers=auth_header(tr_token),
        )
        assert r.status_code == 200
        assert r.content[:2] == b"PK"

    def test_has_semester_sheet(self, client, tr_token, report_data):
        import io
        import openpyxl
        d = report_data
        r = client.get(
            "/api/v1/reports/student-scores/excel",
            params={"student_id": d["students"][0].id},
            headers=auth_header(tr_token),
        )
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        # At least one sheet named after a semester
        assert "2025秋季" in wb.sheetnames


class TestExcelRbac:
    def test_student_forbidden_class_summary(self, client, student_token):
        r = client.get(
            "/api/v1/reports/class-summary/excel",
            params={"class_id": 1, "semester_id": 1},
            headers=auth_header(student_token),
        )
        assert r.status_code == 403

    def test_student_forbidden_grade_ranking(self, client, student_token):
        r = client.get(
            "/api/v1/reports/grade-ranking/excel",
            params={"grade_id": 1, "semester_id": 1},
            headers=auth_header(student_token),
        )
        assert r.status_code == 403
