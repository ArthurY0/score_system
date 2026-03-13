"""
TDD - Tests written FIRST before implementation.

Phase 4: Score Management

User Journeys:
  As a teacher, I want to enter/modify scores for students in my assigned
  course/class/semester, so that the system has accurate grade data.

  As a student, I want to view only my own scores,
  so that my academic data is kept private.

  As an admin, I want to view all scores and audit logs,
  so that I can oversee data integrity and changes.
"""
import io
from datetime import date
import pytest
from tests.conftest import auth_header


# ═══════════════════════════════════════════════════════════
# SHARED FIXTURES
# ═══════════════════════════════════════════════════════════

@pytest.fixture
def semester(db):
    from app.models.semester import Semester
    s = Semester(name="2025-2026学年第一学期", start_date=date(2025, 9, 1), end_date=date(2026, 1, 20), is_active=True)
    db.add(s); db.commit(); db.refresh(s)
    return s


@pytest.fixture
def grade(db):
    from app.models.grade import Grade
    g = Grade(name="高一")
    db.add(g); db.commit(); db.refresh(g)
    return g


@pytest.fixture
def class_(db, grade):
    from app.models.class_ import Class
    c = Class(name="高一(1)班", grade_id=grade.id)
    db.add(c); db.commit(); db.refresh(c)
    return c


@pytest.fixture
def course(db):
    from app.models.course import Course
    c = Course(name="数学")
    db.add(c); db.commit(); db.refresh(c)
    return c


@pytest.fixture
def student(db, class_):
    from app.models.student import Student, Gender
    s = Student(student_no="20250101", name="张三", gender=Gender.MALE, class_id=class_.id)
    db.add(s); db.commit(); db.refresh(s)
    return s


@pytest.fixture
def student2(db, class_):
    from app.models.student import Student, Gender
    s = Student(student_no="20250102", name="李四", gender=Gender.FEMALE, class_id=class_.id)
    db.add(s); db.commit(); db.refresh(s)
    return s


@pytest.fixture
def assignment(db, teacher_user, course, class_, semester):
    from app.models.teacher_course_class import TeacherCourseClass
    a = TeacherCourseClass(teacher_id=teacher_user.id, course_id=course.id, class_id=class_.id, semester_id=semester.id)
    db.add(a); db.commit(); db.refresh(a)
    return a


@pytest.fixture
def score(db, student, course, class_, semester, teacher_user):
    from app.models.score import Score
    s = Score(
        student_id=student.id, course_id=course.id, class_id=class_.id,
        semester_id=semester.id, teacher_id=teacher_user.id, score=85.0,
    )
    db.add(s); db.commit(); db.refresh(s)
    return s


# ═══════════════════════════════════════════════════════════
# SCORE ENTRY TESTS
# ═══════════════════════════════════════════════════════════

class TestScoreEntry:
    def test_teacher_can_submit_score_for_assigned_class(
        self, client, teacher_token, assignment, student, course, class_, semester, teacher_user
    ):
        """Teacher submits a score for a student in their assigned class."""
        r = client.post("/api/v1/scores", json={
            "student_id": student.id,
            "course_id": course.id,
            "class_id": class_.id,
            "semester_id": semester.id,
            "score": 88.5,
        }, headers=auth_header(teacher_token))
        assert r.status_code == 201
        data = r.json()
        assert data["score"] == 88.5
        assert data["student_id"] == student.id

    def test_admin_can_submit_score(self, client, admin_token, student, course, class_, semester, teacher_user):
        r = client.post("/api/v1/scores", json={
            "student_id": student.id, "course_id": course.id,
            "class_id": class_.id, "semester_id": semester.id, "score": 75.0,
        }, headers=auth_header(admin_token))
        assert r.status_code == 201

    def test_teacher_cannot_submit_score_for_unassigned_class(
        self, client, teacher_token, student, course, class_, semester
    ):
        """Teacher without assignment cannot submit scores."""
        r = client.post("/api/v1/scores", json={
            "student_id": student.id, "course_id": course.id,
            "class_id": class_.id, "semester_id": semester.id, "score": 90.0,
        }, headers=auth_header(teacher_token))
        assert r.status_code == 403

    def test_student_cannot_submit_score(self, client, student_token, student, course, class_, semester):
        r = client.post("/api/v1/scores", json={
            "student_id": student.id, "course_id": course.id,
            "class_id": class_.id, "semester_id": semester.id, "score": 99.0,
        }, headers=auth_header(student_token))
        assert r.status_code == 403

    def test_score_above_150_returns_422(self, client, admin_token, student, course, class_, semester):
        r = client.post("/api/v1/scores", json={
            "student_id": student.id, "course_id": course.id,
            "class_id": class_.id, "semester_id": semester.id, "score": 160.0,
        }, headers=auth_header(admin_token))
        assert r.status_code == 422

    def test_score_below_0_returns_422(self, client, admin_token, student, course, class_, semester):
        r = client.post("/api/v1/scores", json={
            "student_id": student.id, "course_id": course.id,
            "class_id": class_.id, "semester_id": semester.id, "score": -1.0,
        }, headers=auth_header(admin_token))
        assert r.status_code == 422

    def test_duplicate_score_returns_409(self, client, admin_token, score, student, course, class_, semester):
        """Cannot submit a second score for same student+course+semester."""
        r = client.post("/api/v1/scores", json={
            "student_id": student.id, "course_id": course.id,
            "class_id": class_.id, "semester_id": semester.id, "score": 70.0,
        }, headers=auth_header(admin_token))
        assert r.status_code == 409

    def test_score_creates_audit_log(self, client, admin_token, student, course, class_, semester, db):
        """Creating a score should create an audit log entry with action=created."""
        r = client.post("/api/v1/scores", json={
            "student_id": student.id, "course_id": course.id,
            "class_id": class_.id, "semester_id": semester.id, "score": 92.0,
        }, headers=auth_header(admin_token))
        assert r.status_code == 201
        score_id = r.json()["id"]
        from app.models.score import ScoreAuditLog
        logs = db.query(ScoreAuditLog).filter_by(score_id=score_id).all()
        assert len(logs) == 1
        assert logs[0].action == "created"
        assert logs[0].new_score == 92.0


# ═══════════════════════════════════════════════════════════
# SCORE QUERY TESTS
# ═══════════════════════════════════════════════════════════

class TestScoreQuery:
    def test_admin_can_list_all_scores(self, client, admin_token, score):
        r = client.get("/api/v1/scores", headers=auth_header(admin_token))
        assert r.status_code == 200
        assert r.json()["total"] >= 1

    def test_filter_scores_by_class_and_semester(self, client, admin_token, score, class_, semester):
        r = client.get(
            f"/api/v1/scores?class_id={class_.id}&semester_id={semester.id}",
            headers=auth_header(admin_token),
        )
        assert r.status_code == 200
        items = r.json()["items"]
        assert all(s["class_id"] == class_.id for s in items)

    def test_filter_scores_by_course(self, client, admin_token, score, course):
        r = client.get(f"/api/v1/scores?course_id={course.id}", headers=auth_header(admin_token))
        assert r.status_code == 200

    def test_teacher_can_list_scores_for_assigned_class(self, client, teacher_token, assignment, score):
        r = client.get("/api/v1/scores", headers=auth_header(teacher_token))
        assert r.status_code == 200

    def test_student_can_view_own_scores(self, client, student_token, student_user, score, db):
        """Student can only see their own scores via /scores/my."""
        from app.models.student import Student
        stu = db.query(Student).filter_by(student_no="20250101").first()
        stu.user_id = student_user.id
        db.commit()

        r = client.get("/api/v1/scores/my", headers=auth_header(student_token))
        assert r.status_code == 200
        for item in r.json()["items"]:
            assert item["student_id"] == stu.id

    def test_student_cannot_access_all_scores(self, client, student_token):
        r = client.get("/api/v1/scores", headers=auth_header(student_token))
        assert r.status_code == 403

    def test_admin_can_get_single_score(self, client, admin_token, score):
        r = client.get(f"/api/v1/scores/{score.id}", headers=auth_header(admin_token))
        assert r.status_code == 200
        assert r.json()["score"] == 85.0

    def test_get_nonexistent_score_returns_404(self, client, admin_token):
        assert client.get("/api/v1/scores/99999", headers=auth_header(admin_token)).status_code == 404


# ═══════════════════════════════════════════════════════════
# SCORE UPDATE TESTS (with audit log)
# ═══════════════════════════════════════════════════════════

class TestScoreUpdate:
    def test_teacher_can_update_own_score(self, client, teacher_token, assignment, score):
        r = client.put(f"/api/v1/scores/{score.id}",
            json={"score": 90.0, "reason": "录入错误更正"},
            headers=auth_header(teacher_token))
        assert r.status_code == 200
        assert r.json()["score"] == 90.0

    def test_score_update_creates_audit_log(self, client, admin_token, score, db):
        client.put(f"/api/v1/scores/{score.id}",
            json={"score": 95.0, "reason": "修改分数"},
            headers=auth_header(admin_token))
        from app.models.score import ScoreAuditLog
        logs = db.query(ScoreAuditLog).filter_by(score_id=score.id).all()
        update_log = next((l for l in logs if l.action == "updated"), None)
        assert update_log is not None
        assert update_log.old_score == 85.0
        assert update_log.new_score == 95.0
        assert update_log.reason == "修改分数"

    def test_teacher_cannot_update_another_teachers_score(
        self, client, db, score, semester, course, class_
    ):
        """A teacher not assigned to the course/class cannot modify the score."""
        from app.models.user import User, UserRole
        from app.core.security import hash_password
        other_teacher = User(
            username="teacher2", hashed_password=hash_password("pass1234"),
            name="其他老师", role=UserRole.TEACHER, is_active=True,
        )
        db.add(other_teacher); db.commit(); db.refresh(other_teacher)
        r_login = client.post("/api/v1/auth/login",
            data={"username": "teacher2", "password": "pass1234"},
            headers={"Content-Type": "application/x-www-form-urlencoded"})
        other_token = r_login.json()["access_token"]

        r = client.put(f"/api/v1/scores/{score.id}",
            json={"score": 99.0, "reason": "hack"},
            headers=auth_header(other_token))
        assert r.status_code == 403

    def test_update_score_above_150_returns_422(self, client, admin_token, score):
        r = client.put(f"/api/v1/scores/{score.id}",
            json={"score": 200.0}, headers=auth_header(admin_token))
        assert r.status_code == 422

    def test_admin_can_delete_score(self, client, admin_token, score):
        assert client.delete(f"/api/v1/scores/{score.id}", headers=auth_header(admin_token)).status_code == 204

    def test_teacher_cannot_delete_score(self, client, teacher_token, assignment, score):
        assert client.delete(f"/api/v1/scores/{score.id}", headers=auth_header(teacher_token)).status_code == 403


# ═══════════════════════════════════════════════════════════
# BATCH SCORE ENTRY
# ═══════════════════════════════════════════════════════════

class TestBatchScoreEntry:
    def test_admin_can_batch_create_scores(self, client, admin_token, student, student2, course, class_, semester, teacher_user):
        r = client.post("/api/v1/scores/batch", json={
            "class_id": class_.id, "course_id": course.id,
            "semester_id": semester.id, "teacher_id": teacher_user.id,
            "entries": [
                {"student_id": student.id, "score": 88.0},
                {"student_id": student2.id, "score": 76.5},
            ],
        }, headers=auth_header(admin_token))
        assert r.status_code == 200
        data = r.json()
        assert data["created"] == 2
        assert data["updated"] == 0
        assert data["errors"] == []

    def test_batch_upsert_updates_existing_scores(self, client, admin_token, score, student, student2, course, class_, semester, teacher_user):
        """Batch entry on an existing score should update it, not create a duplicate."""
        r = client.post("/api/v1/scores/batch", json={
            "class_id": class_.id, "course_id": course.id,
            "semester_id": semester.id, "teacher_id": teacher_user.id,
            "entries": [
                {"student_id": student.id, "score": 91.0},   # existing → update
                {"student_id": student2.id, "score": 82.0},  # new → create
            ],
        }, headers=auth_header(admin_token))
        assert r.status_code == 200
        assert r.json()["created"] == 1
        assert r.json()["updated"] == 1

    def test_teacher_can_batch_submit_for_assigned_class(
        self, client, teacher_token, assignment, student, course, class_, semester, teacher_user
    ):
        r = client.post("/api/v1/scores/batch", json={
            "class_id": class_.id, "course_id": course.id,
            "semester_id": semester.id, "teacher_id": teacher_user.id,
            "entries": [{"student_id": student.id, "score": 77.0}],
        }, headers=auth_header(teacher_token))
        assert r.status_code == 200

    def test_teacher_cannot_batch_submit_for_unassigned_class(
        self, client, teacher_token, student, course, class_, semester, teacher_user
    ):
        r = client.post("/api/v1/scores/batch", json={
            "class_id": class_.id, "course_id": course.id,
            "semester_id": semester.id, "teacher_id": teacher_user.id,
            "entries": [{"student_id": student.id, "score": 77.0}],
        }, headers=auth_header(teacher_token))
        assert r.status_code == 403


# ═══════════════════════════════════════════════════════════
# EXCEL IMPORT
# ═══════════════════════════════════════════════════════════

class TestScoreExcelImport:
    def test_admin_can_import_scores_via_excel(self, client, admin_token, student, student2, course, class_, semester, teacher_user):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["学号", "分数"])
        ws.append(["20250101", 93.0])
        ws.append(["20250102", 87.5])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)

        r = client.post(
            f"/api/v1/scores/import?class_id={class_.id}&course_id={course.id}"
            f"&semester_id={semester.id}&teacher_id={teacher_user.id}",
            files={"file": ("scores.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_header(admin_token),
        )
        assert r.status_code == 200
        data = r.json()
        assert data["created"] == 2
        assert data["errors"] == []

    def test_excel_import_reports_unknown_student_no(self, client, admin_token, course, class_, semester, teacher_user):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["学号", "分数"])
        ws.append(["UNKNOWN999", 80.0])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)

        r = client.post(
            f"/api/v1/scores/import?class_id={class_.id}&course_id={course.id}"
            f"&semester_id={semester.id}&teacher_id={teacher_user.id}",
            files={"file": ("scores.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_header(admin_token),
        )
        assert r.status_code == 200
        assert len(r.json()["errors"]) == 1

    def test_excel_import_invalid_score_reports_error(self, client, admin_token, student, course, class_, semester, teacher_user):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["学号", "分数"])
        ws.append(["20250101", "abc"])  # invalid score
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)

        r = client.post(
            f"/api/v1/scores/import?class_id={class_.id}&course_id={course.id}"
            f"&semester_id={semester.id}&teacher_id={teacher_user.id}",
            files={"file": ("scores.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=auth_header(admin_token),
        )
        assert r.status_code == 200
        assert len(r.json()["errors"]) >= 1


# ═══════════════════════════════════════════════════════════
# AUDIT LOG TESTS
# ═══════════════════════════════════════════════════════════

class TestAuditLog:
    def test_admin_can_view_audit_logs(self, client, admin_token, score):
        r = client.get(f"/api/v1/scores/{score.id}/audit-logs", headers=auth_header(admin_token))
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_teacher_can_view_audit_logs_for_own_score(self, client, teacher_token, assignment, score):
        r = client.get(f"/api/v1/scores/{score.id}/audit-logs", headers=auth_header(teacher_token))
        assert r.status_code == 200

    def test_student_cannot_view_audit_logs(self, client, student_token, score):
        r = client.get(f"/api/v1/scores/{score.id}/audit-logs", headers=auth_header(student_token))
        assert r.status_code == 403
